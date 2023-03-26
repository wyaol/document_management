from zipfile import ZipFile

import docx
import openpyxl
import copy
from io import BytesIO


def read_doc_to_data_set(bytes_file):
    # 打开文档
    byte_stream = BytesIO(bytes_file)
    doc = docx.Document(byte_stream)
    data_sets = []
    # 输出每个段落的文本内容
    for para in doc.paragraphs:
        if para.text != '':
            data_sets.append(para.text)
    return data_sets


def get_actual_length(data_array):
    res = 0
    for item in data_array:
        if '' != item:
            res += 1
    return res


def get_max_length(data_array_list):
    res = 0
    for item in data_array_list:
        res = max(get_actual_length(item), res)
    return res


def classify_titles(titles):
    res = []
    print_array = [''] * 10
    for i in range(1, len(titles)):
        if '第' in titles[i]:
            print_array = [''] * 10
            print_array[0] = titles[i]
        else:
            tittle_level = len(titles[i].split('.'))
            print_array[tittle_level - 1] = titles[i]

            # 如果下一个标题不是该标题第子标题 或者这是最后一个标题，说明需要将该标题录入
            if i == len(titles) - 1 or len(titles[i + 1].split('.')) <= tittle_level:
                array_copy = copy.deepcopy(print_array)
                res.append(array_copy)
    return res


def get_excel_workbook(data_sets, file_name):
    number_map = ['一', '二', '三', '四', '五', '六']
    # 打开工作簿（Excel文件）
    workbook = openpyxl.Workbook()
    # 获取工作簿中的所有工作表（sheet）
    sheets = workbook.sheetnames
    # 选择第一个工作表
    worksheet = workbook[sheets[0]]
    # 填充章节部分
    data_sets_list = classify_titles(data_sets)
    column_number = get_max_length(data_sets_list)
    for i_ in range(column_number):
        worksheet.cell(row=1, column=i_ + 1, value='%s级标题' % number_map[i_])
    print(data_sets_list)
    for i in range(len(data_sets_list)):
        for j in range(len(data_sets_list[i])):
            worksheet.cell(row=i + 2, column=j + 1, value=data_sets_list[i][j])
    # 添加题目
    title = file_name.split('.')[0]
    worksheet.cell(row=1, column=column_number + 1, value='题目')
    for i in range(len(data_sets_list)):
        worksheet.cell(row=2 + i, column=column_number + 1, value=title)
    # 保存工作簿
    return {
        'title': title,
        'workbook': workbook
    }


def get_zip_output(workbook_obj_list):
    # {
    #     'title': title,
    #     'workbook': workbook
    # }[]
    # 创建ZipFile对象，并将Workbook对象列表中的所有对象添加到zip文件中
    output = BytesIO()
    with ZipFile(output, 'w') as zip_file:
        for workbook_obj in workbook_obj_list:
            workbook_obj['workbook'].save(f'output/%s.xlsx' % workbook_obj['title'])
            zip_file.write(f'output/%s.xlsx' % workbook_obj['title'])
    return output
